
#from re import A
import pandas as pd
import numpy as np
import string
from openpyxl import load_workbook
import string

archivo_excel = pd.read_excel('C:/Users/salva/Downloads/SUS_deploymentPiloto (respuestas).xlsx')
prueba=archivo_excel[['Pienso que me gustaría usar este sistema con frecuencia. ', 'Me parece que el sistema es complejo sin necesidad de serlo. ','Me ha parecido que el sistema era fácil de usar. ','Creo que necesitaría el apoyo de algún técnico para poder usar el sistema. ','Me ha parecido que las funciones del sistema estaban bien integradas. ','Me ha parecido que el sistema tenía demasiadas inconsistencias. ', 'Me parece que la mayoría de las personas van a aprender a usar el sistema muy rápido. ','Me ha parecido que el sistema era muy engorroso de usar. ','Me he sentido muy seguro/a al usar el sistema. ','Necesitaría aprender muchas cosas antes de poder empezar a usar el sistema. ', 'Marca temporal', 'Código Pharaon']]

prueba.to_excel('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/SUS/prueba_sus.xlsx', sheet_name='Worksheet')
wb = load_workbook('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/SUS/prueba_sus.xlsx')
wb.iso_dates=True
ws=wb['Worksheet']
ws.delete_cols(1)
ws.insert_cols(11,12)
ws.insert_cols(24,1)

min_col=wb.active.min_column
min_fila=wb.active.min_row
max_col=wb.active.max_column
max_fila=wb.active.max_row

#wb.save('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/SUS/prueba_sus.xlsx')

abecedario=list(string.ascii_uppercase) #letras abecedario en mayusculas
abecedario_excel=abecedario[0:max_col] #desde A hasta Y, q es dde estan los datos

cont=0
for i in abecedario_excel[0:20]:
    cont=cont+1
    if cont<10:
        ws[f'{i}1']=f'answer_0{cont}'
    else:
        ws[f'{i}1']=f'answer_{cont}'

ws['U1']='type'
ws['V1']='phase'
ws['W1']='date_filling'
ws['X1']='score'
ws['Y1']='recruiment_id'

wb.save('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/SUS/prueba_sus.xlsx')

score=np.zeros(max_fila-1)

for i in range(max_fila-1):
    for j in abecedario_excel[0:10]:
        if ws[f'{j}{i+2}'].value == 1:
            if j=='A' or j=='C' or j=='E' or j=='G' or j=='I':
                score[i]=score[i]+0
            if j=='B' or j=='D' or j=='F' or j=='H' or j=='J':
                score[i]=score[i]+4
        if ws[f'{j}{i+2}'].value == 2:
            if j=='A' or j=='C' or j=='E' or j=='G' or j=='I':
                score[i]=score[i]+1
            if j=='B' or j=='D' or j=='F' or j=='H' or j=='J':
                score[i]=score[i]+3
        if ws[f'{j}{i+2}'].value == 3:
            if j=='A' or j=='C' or j=='E' or j=='G' or j=='I':
                score[i]=score[i]+2
            if j=='B' or j=='D' or j=='F' or j=='H' or j=='J':
                score[i]=score[i]+2
        if ws[f'{j}{i+2}'].value == 4:
            if j=='A' or j=='C' or j=='E' or j=='G' or j=='I':
                score[i]=score[i]+3
            if j=='B' or j=='D' or j=='F' or j=='H' or j=='J':
                score[i]=score[i]+1
        if ws[f'{j}{i+2}'].value == 5:
            if j=='A' or j=='C' or j=='E' or j=='G' or j=='I':
                score[i]=score[i]+4
            if j=='B' or j=='D' or j=='F' or j=='H' or j=='J':
                score[i]=score[i]+0


for i in range(max_fila-1):
    ws[f'U{i+2}'].value = 'sus'
    ws[f'V{i+2}'].value = 'Baseline'
    ws[f'X{i+2}'].value = score[i]*2.5
    year_extension = str(ws[f'W{i+2}'].value).split(' ')[0]
    ws[f'W{i+2}'].value = (f'{year_extension}')
    #arg1=ws[f'Y{i+2}'].value.split('_')[0]
    #arg2=ws[f'Y{i+2}'].value.split('_')[1]
    #arg3=ws[f'Y{i+2}'].value.split('_')[2]
    ws[f'Y{i+2}'].value = str(ws[f'Y{i+2}'].value).lower()
    

wb.save('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/SUS/final_sus.xlsx')