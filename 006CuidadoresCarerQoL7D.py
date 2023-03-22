
#from re import A
import pandas as pd
import numpy as np
import string
from openpyxl import load_workbook
import string

archivo_excel = pd.read_excel('C:/Users/salva/Downloads/Cuestionario-difusiónWhatsapp-CuidadoresInformales(UCLAyCareQoL7D) (respuestas).xlsx')
prueba=archivo_excel[['Por favor, valore cada una de las siguientes afirmaciones con "No", "Un poco" o  "Mucho" de acuerdo con la que más se adecúe a su situación como cuidador. [Me satisface desempeñar mis tareas de cuidador]', 'Por favor, valore cada una de las siguientes afirmaciones con "No", "Un poco" o  "Mucho" de acuerdo con la que más se adecúe a su situación como cuidador. [Tengo problemas de relación con la persona a la que cuido (ej.: es muy exigente, o se comporta de manera distinta, problemas de comunicación, etc.)]', 'Por favor, valore cada una de las siguientes afirmaciones con "No", "Un poco" o  "Mucho" de acuerdo con la que más se adecúe a su situación como cuidador. [Tengo problemas con mi propia salud mental (ej.: estrés, miedo, pesimismo, depresión, preocupación por el futuro).]', 'Por favor, valore cada una de las siguientes afirmaciones con "No", "Un poco" o  "Mucho" de acuerdo con la que más se adecúe a su situación como cuidador. [Tengo problemas para compaginar mis tareas con mis actividades diarias (Ej.: tareas domésticas, trabajo, estudios, familia, ocio).]', 'Por favor, valore cada una de las siguientes afirmaciones con "No", "Un poco" o  "Mucho" de acuerdo con la que más se adecúe a su situación como cuidador. [Tengo problemas financieros debido a mi labor como cuidador/a.]', 'Por favor, valore cada una de las siguientes afirmaciones con "No", "Un poco" o  "Mucho" de acuerdo con la que más se adecúe a su situación como cuidador. [Recibo ayuda de otros para mis tareas como cuidador/a (Ej.: de familiares, amigos, vecinos, conocidos), cuando los necesito.]', 'Por favor, valore cada una de las siguientes afirmaciones con "No", "Un poco" o  "Mucho" de acuerdo con la que más se adecúe a su situación como cuidador. [Tengo problemas con mi propia salud física (Ej.: enfermo con más frecuencia, cansancio, estrés físico, etc.).]', 'Marca temporal', 'Código Pharaon (opcional)']]
#print(prueba)
prueba.to_excel('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/Cuidadores Informales/prueba_carerqol7d.xlsx', sheet_name='Worksheet')
wb = load_workbook('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/Cuidadores Informales/prueba_carerqol7d.xlsx')
wb.iso_dates=True
ws=wb['Worksheet']
ws.delete_cols(1)
ws.insert_cols(8,15)
ws.insert_cols(24,1)

min_col=wb.active.min_column
min_fila=wb.active.min_row
max_col=wb.active.max_column
max_fila=wb.active.max_row

wb.save('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/Cuidadores Informales/prueba_carerqol7d.xlsx')

abecedario=list(string.ascii_uppercase) #letras abecedario en mayusculas
abecedario_excel=abecedario[0:max_col] #desde A hasta Y, q es dde estan los datos

cont=0
for i in abecedario_excel[0:20]:
    cont=cont+1
    if cont<10:
        ws[f'{i}1']=f'answer_0{cont}'
    else:
        ws[f'{i}1']=f'answer_{cont}'

ws['U1']='type'
ws['V1']='phase'
ws['W1']='date_filling'
ws['X1']='score'
ws['Y1']='recruiment_id'

wb.save('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/Cuidadores Informales/prueba_carerqol7d.xlsx')

score=np.zeros(max_fila-1)

for i in range(max_fila-1):
    for j in abecedario_excel[0:7]:
        if j=='A':
            if ws[f'{j}{i+2}'].value == 'No':
                ws[f'{j}{i+2}'].value = 0
            if ws[f'{j}{i+2}'].value == 'Un poco':
                ws[f'{j}{i+2}'].value = 1
                score[i]=score[i]+11.9
            if ws[f'{j}{i+2}'].value == 'Mucho':
                ws[f'{j}{i+2}'].value = 2
                score[i]=score[i]+15.1
        if j=='B':
            if ws[f'{j}{i+2}'].value == 'No':
                ws[f'{j}{i+2}'].value = 2
                score[i]=score[i]+17.8
            if ws[f'{j}{i+2}'].value == 'Un poco':
                ws[f'{j}{i+2}'].value = 1
                score[i]=score[i]+10.6
            if ws[f'{j}{i+2}'].value == 'Mucho':
                ws[f'{j}{i+2}'].value = 0
        if j=='C':
            if ws[f'{j}{i+2}'].value == 'No':
                ws[f'{j}{i+2}'].value = 2
                score[i]=score[i]+13.5
            if ws[f'{j}{i+2}'].value == 'Un poco':
                ws[f'{j}{i+2}'].value = 1
                score[i]=score[i]+12.7
            if ws[f'{j}{i+2}'].value == 'Mucho':
                ws[f'{j}{i+2}'].value = 0
        if j=='D':
            if ws[f'{j}{i+2}'].value == 'No':
                ws[f'{j}{i+2}'].value = 2
                score[i]=score[i]+9.2
            if ws[f'{j}{i+2}'].value == 'Un poco':
                ws[f'{j}{i+2}'].value = 1
                score[i]=score[i]+6
            if ws[f'{j}{i+2}'].value == 'Mucho':
                ws[f'{j}{i+2}'].value = 0
        if j=='E':
            if ws[f'{j}{i+2}'].value == 'No':
                ws[f'{j}{i+2}'].value = 2
                score[i]=score[i]+20.3
            if ws[f'{j}{i+2}'].value == 'Un poco':
                ws[f'{j}{i+2}'].value = 1
                score[i]=score[i]+13.4
            if ws[f'{j}{i+2}'].value == 'Mucho':
                ws[f'{j}{i+2}'].value = 0
        if j=='F':
            if ws[f'{j}{i+2}'].value == 'No':
                ws[f'{j}{i+2}'].value = 0
            if ws[f'{j}{i+2}'].value == 'Un poco':
                ws[f'{j}{i+2}'].value = 1
                score[i]=score[i]+6.6
            if ws[f'{j}{i+2}'].value == 'Mucho':
                ws[f'{j}{i+2}'].value = 2
                score[i]=score[i]+9.1
        if j=='G':
            if ws[f'{j}{i+2}'].value == 'No':
                ws[f'{j}{i+2}'].value = 2
                score[i]=score[i]+15.1
            if ws[f'{j}{i+2}'].value == 'Un poco':
                ws[f'{j}{i+2}'].value = 1
                score[i]=score[i]+13.2
            if ws[f'{j}{i+2}'].value == 'Mucho':
                ws[f'{j}{i+2}'].value = 0



for i in range(max_fila-1):
    ws[f'U{i+2}'].value = 'carerqol-7d'
    ws[f'V{i+2}'].value = 'Baseline'
    ws[f'X{i+2}'].value = score[i]
    year_extension = str(ws[f'W{i+2}'].value).split(' ')[0]
    ws[f'W{i+2}'].value = (f'{year_extension}')
    arg1=ws[f'Y{i+2}'].value.split('_')[0]
    arg2=ws[f'Y{i+2}'].value.split('_')[1]
    arg3=ws[f'Y{i+2}'].value.split('_')[2]
    ws[f'Y{i+2}'].value = (f'{arg1}_{arg2}_{arg3}_ic').lower()
    

wb.save('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/Cuidadores Informales/final_carerqol7d.xlsx')