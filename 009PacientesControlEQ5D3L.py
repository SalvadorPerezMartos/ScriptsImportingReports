
#from re import A
import pandas as pd
import numpy as np
import string
from openpyxl import load_workbook
import string

archivo_excel = pd.read_excel('C:/Users/salva/Downloads/Cuestionario-difusiónWhatsapp-Pacientes(EQ-5D-3L y UCLA) (respuestas) (2).xlsx')
prueba=archivo_excel[['Movilidad', 'Cuidado personal', 'Actividades cotidianas (trabajar, estudiar, tareas domésticas, actividades familiares o durante el tiempo libre)', 'Dolor/Malestar', 'Ansiedad/Depresión', 'Valore, en una escala de 0 a 100, el estado de salud del paciente en el día de HOY.', 'Marca temporal', 'Código Pharaon (opcional)']]
#print(prueba)
prueba.to_excel('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/Pacientes CONTROL/prueba_eq-5d-3l.xlsx', sheet_name='Worksheet')

wb = load_workbook('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/Pacientes CONTROL/prueba_eq-5d-3l.xlsx')
wb.iso_dates=True
ws=wb['Worksheet']
ws.delete_cols(1)
ws.insert_cols(7,16)
ws.insert_cols(24)
min_col=wb.active.min_column
min_fila=wb.active.min_row
max_col=wb.active.max_column
max_fila=wb.active.max_row

wb.save('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/Pacientes CONTROL/prueba_eq-5d-3l.xlsx')

abecedario=list(string.ascii_uppercase) #letras abecedario en mayusculas
abecedario_excel=abecedario[0:max_col] #desde A hasta Y, q es dde estan los datos

cont=0
for i in abecedario_excel[0:20]:
    cont=cont+1
    if cont<10:
        ws[f'{i}1']=f'answer_0{cont}'
    else:
        ws[f'{i}1']=f'answer_{cont}'

ws['U1']='type'
ws['V1']='phase'
ws['W1']='date_filling'
ws['X1']='score'
ws['Y1']='recruiment_id'

numberOf2=np.zeros(max_fila-1)
numberOf3=np.zeros(max_fila-1)
score=np.ones(max_fila-1)


for i in range(max_fila-1):

    #movilidad
    if ws[f'A{i+2}'].value == 'El paciente no tiene problemas para caminar':
        ws[f'A{i+2}'].value = 1
    if ws[f'A{i+2}'].value == 'El paciente tiene algunos problemas para caminar':
        ws[f'A{i+2}'].value = 2
        numberOf2[i]=numberOf2[i]+1
        score[i]=score[i]-0.0659
    if ws[f'A{i+2}'].value == 'El paciente tiene que permanecer en la cama':
        ws[f'A{i+2}'].value = 3
        numberOf3[i]=numberOf3[i]+1
        score[i]=score[i]-0.1829

    #cuidado personal
    if ws[f'B{i+2}'].value == 'El paciente no tiene problemas para lavarse y/o vestirse':
        ws[f'B{i+2}'].value = 1
    if ws[f'B{i+2}'].value == 'El paciente tiene algunos problemas para lavarse y/o vestirse':
        ws[f'B{i+2}'].value = 2
        numberOf2[i]=numberOf2[i]+1
        score[i]=score[i]-0.1173
    if ws[f'B{i+2}'].value == 'El paciente es incapaz de lavarse o vestirse por sí solo':
        ws[f'B{i+2}'].value = 3
        numberOf3[i]=numberOf3[i]+1
        score[i]=score[i]-0.1559

    #actividades cotidianas
    if ws[f'C{i+2}'].value == 'El paciente no tiene problemas para realizar sus actividades cotidianas':
        ws[f'C{i+2}'].value = 1
    if ws[f'C{i+2}'].value == 'El paciente tiene algunos problemas para realizar sus actividades cotidianas':
        ws[f'C{i+2}'].value = 2
        numberOf2[i]=numberOf2[i]+1
        score[i]=score[i]-0.0264
    if ws[f'C{i+2}'].value == 'El paciente es incapaz de realizar sus actividades cotidianas':
        ws[f'C{i+2}'].value = 3
        numberOf3[i]=numberOf3[i]+1
        score[i]=score[i]-0.086
    
    #dolor/malestar
    if ws[f'D{i+2}'].value == 'El paciente no tiene dolor ni malestar':
        ws[f'D{i+2}'].value = 1
    if ws[f'D{i+2}'].value == 'El paciente tiene dolores/malestares moderados':
        ws[f'D{i+2}'].value = 2
        numberOf2[i]=numberOf2[i]+1
        score[i]=score[i]-0.093
    if ws[f'D{i+2}'].value == 'El paciente tiene mucho dolor o malestar':
        ws[f'D{i+2}'].value = 3
        numberOf3[i]=numberOf3[i]+1
        score[i]=score[i]-0.1637
    
    #ansiedad/depresion
    if ws[f'E{i+2}'].value == 'El paciente no está ansioso o deprimido':
        ws[f'E{i+2}'].value = 1
    if ws[f'E{i+2}'].value == 'El paciente está moderadamente ansioso o deprimido':
        ws[f'E{i+2}'].value = 2
        numberOf2[i]=numberOf2[i]+1
        score[i]=score[i]-0.0891
    if ws[f'E{i+2}'].value == 'El pacientes está muy ansioso o deprimido':
        ws[f'E{i+2}'].value = 3
        numberOf3[i]=numberOf3[i]+1
        score[i]=score[i]-0.129



for i in range(max_fila-1):
    if numberOf2[i]>0 and numberOf3[i] == 0:
        score[i]=score[i]-0.1279
    if numberOf3[i]>0:
        score[i]=score[i]-0.2279


for i in range(max_fila-1):
    ws[f'U{i+2}'].value = 'eq-5d-3l'
    ws[f'V{i+2}'].value = 'Baseline'
    ws[f'X{i+2}'].value = score[i]
    year_extension = str(ws[f'W{i+2}'].value).split(' ')[0]
    ws[f'W{i+2}'].value = (f'{year_extension}')
    arg1=ws[f'Y{i+2}'].value.split('_')[0]
    arg2=ws[f'Y{i+2}'].value.split('_')[1]
    arg3=ws[f'Y{i+2}'].value.split('_')[2]
    ws[f'Y{i+2}'].value = (f'{arg1}_{arg2}_{arg3}_oa').lower()


wb.save('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/Pacientes CONTROL/final_eq-5d-3l.xlsx')