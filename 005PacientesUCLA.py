
#from re import A
import pandas as pd
import numpy as np
import string
from openpyxl import load_workbook
import string

archivo_excel = pd.read_excel('C:/Users/salva/Downloads/Cuestionario-difusiónWhatsapp-Pacientes(EQ-5D-3L y UCLA) (respuestas) (1).xlsx')
prueba=archivo_excel[['Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia siente el paciente que está en sintonía (se lleva bien) con la gente que le rodea?]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia siente el paciente que le falta compañía?]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia siente el paciente que no tiene a nadie con quién pueda contar?]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia el paciente se siente solo/a? ]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia el paciente se siente parte de un grupo de amigos/as?]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia siente el paciente que tiene muchas cosas en común con la gente que le rodea? ]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia siente el paciente que no tiene confianza con nadie? ]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia siente el paciente que sus intereses e ideas no son compartidos por las personas que le rodean? ]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia siente el paciente que es una persona abierta (extrovertida) y amable? ]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia el paciente se siente cercano/a de algunas personas? ]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia el paciente se siente excluido/a, olvidado/a por los demás? ]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia el paciente siente que sus relaciones sociales son superficiales?]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia piensa el paciente que realmente nadie le conoce bien? ]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia el paciente se siente aislado/a de los demás? ]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia siente el paciente que puede encontrar compañía cuando lo desea?]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia siente el paciente que hay personas que realmente le comprenden? ]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia el paciente se siente infeliz de estar tan aislado/a? ]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia siente el paciente que la gente está a su alrededor pero no siente que esté con él? ]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia siente el paciente que hay personas con las que puede charlar y comunicarse? ]', 'Las siguientes afirmaciones describen cómo se siente a veces la gente. Para cada afirmación, indique con qué frecuencia considera que usted, como paciente, se siente de la manera descrita, marcando el espacio correspondiente. He aquí un ejemplo: ¿Con qué frecuencia se siente feliz? Si nunca se siente feliz, marcaría "nunca"; si siempre se siente feliz, marcaría "siempre". [¿Con qué frecuencia siente el paciente que hay personas a las que puede recurrir?]', 'Marca temporal', 'Código Pharaon (opcional)']]
#print(prueba)
prueba.to_excel('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/Pacientes/prueba_ucla.xlsx', sheet_name='Worksheet')

wb = load_workbook('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/Pacientes/prueba_ucla.xlsx')
wb.iso_dates=True
ws=wb['Worksheet']
ws.delete_cols(1)
ws.insert_cols(21,2)
ws.insert_cols(24,1)

min_col=wb.active.min_column
min_fila=wb.active.min_row
max_col=wb.active.max_column
max_fila=wb.active.max_row

wb.save('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/Pacientes/prueba_ucla.xlsx')

abecedario=list(string.ascii_uppercase) #letras abecedario en mayusculas
abecedario_excel=abecedario[0:max_col] #desde A hasta Y, q es dde estan los datos

cont=0
for i in abecedario_excel[0:20]:
    cont=cont+1
    if cont<10:
        ws[f'{i}1']=f'answer_0{cont}'
    else:
        ws[f'{i}1']=f'answer_{cont}'

ws['U1']='type'
ws['V1']='phase'
ws['W1']='date_filling'
ws['X1']='score'
ws['Y1']='recruiment_id'

score=np.zeros(max_fila-1)

for i in range(max_fila-1):
    for j in abecedario_excel[0:20]:
        if ws[f'{j}{i+2}'].value == 'Siempre':
            ws[f'{j}{i+2}'].value = 4
            if j=='B' or j=='C' or j=='D' or j=='G' or j=='H' or j=='K' or j=='L' or j=='M' or j=='N' or j=='Q' or j=='R':
                score[i]=score[i]+4
            if j=='A' or j=='E' or j=='F' or j=='I' or j=='J' or j=='O' or j=='P' or j=='S' or j=='T':
                score[i]=score[i]+1
        if ws[f'{j}{i+2}'].value == 'A veces':
            ws[f'{j}{i+2}'].value = 3
            if j=='B' or j=='C' or j=='D' or j=='G' or j=='H' or j=='K' or j=='L' or j=='M' or j=='N' or j=='Q' or j=='R':
                score[i]=score[i]+3
            if j=='A' or j=='E' or j=='F' or j=='I' or j=='J' or j=='O' or j=='P' or j=='S' or j=='T':
                score[i]=score[i]+2
        if ws[f'{j}{i+2}'].value == 'Rara vez':
            ws[f'{j}{i+2}'].value = 2
            if j=='B' or j=='C' or j=='D' or j=='G' or j=='H' or j=='K' or j=='L' or j=='M' or j=='N' or j=='Q' or j=='R':
                score[i]=score[i]+2
            if j=='A' or j=='E' or j=='F' or j=='I' or j=='J' or j=='O' or j=='P' or j=='S' or j=='T':
                score[i]=score[i]+3
        if ws[f'{j}{i+2}'].value == 'Nunca':
            ws[f'{j}{i+2}'].value = 1
            if j=='B' or j=='C' or j=='D' or j=='G' or j=='H' or j=='K' or j=='L' or j=='M' or j=='N' or j=='Q' or j=='R':
                score[i]=score[i]+1
            if j=='A' or j=='E' or j=='F' or j=='I' or j=='J' or j=='O' or j=='P' or j=='S' or j=='T':
                score[i]=score[i]+4


for i in range(max_fila-1):
    ws[f'U{i+2}'].value = 'ucla'
    ws[f'V{i+2}'].value = 'Baseline'
    ws[f'X{i+2}'].value = score[i]
    year_extension = str(ws[f'W{i+2}'].value).split(' ')[0]
    ws[f'W{i+2}'].value = (f'{year_extension}')
    #arg1=ws[f'Y{i+2}'].value.split('_')[0]
    #arg2=ws[f'Y{i+2}'].value.split('_')[1]
    #arg3=ws[f'Y{i+2}'].value.split('_')[2]
    #ws[f'Y{i+2}'].value = (f'{arg1}_{arg2}_{arg3}_ic').lower()
    ws[f'Y{i+2}'].value = ws[f'Y{i+2}'].value.lower()
    

wb.save('C:/Users/salva/OneDrive/Escritorio/Imports/Reports/Prueba Scripts/Pacientes/final_ucla.xlsx')